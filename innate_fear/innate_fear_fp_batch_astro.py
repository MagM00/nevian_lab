# -*- coding: utf-8 -*-
"""
Created on Mon Jul 24 10:47:41 2023

@author: huang
"""

import os
import h5py
import tkinter
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy import signal
from scipy.stats import zscore
from data_import import import_ppd
from scipy.signal import savgol_filter
from openpyxl import load_workbook

# Set the PPD files directory
#ppd_files_dir = r'H:\fp_test\innate_fear\grabda'
#ppd_files_dir = r'H:\fp_test\innate_fear\gcamp'
#ppd_files_dir = r'H:\fp_test\innate_fear\grab5ht'
ppd_files_dir = r'H:\fp_test\innate_fear\astro'
#ppd_files_dir = r'H:\fp_test\innate_fear\astro\test_jun'

# Get a list of all PPD files in the directory
ppd_files = [f for f in os.listdir(ppd_files_dir) if f.endswith('.ppd')]

timing_data = {
        '0001': 
            {'Stim': [0, 0, 1, 0, 1, 1, 1, 1, 0, 0, 0, 1, 1, 1, 0, 1, 1, 0, 0, 1, 1, 0, 0, 0, 1, 1, 0, 0, 1, 0], 
             'Time': ['00:29', '01:27', '01:58', '02:28', '03:22', '03:56', '04:35', '05:23', '06:23', '07:22', 
                      '08:09', '08:55', '09:38', '10:12', '10:53', '11:32', '12:14', '12:59', '13:33', '14:10', 
                      '15:02', '15:34', '16:08', '16:41', '17:37', '18:08', '18:50', '19:21', '19:59', '20:58']}, 
        '1191': 
            {'Stim': [0, 0, 1, 1, 1, 0, 0, 1, 1, 1, 0, 1, 1, 0, 1, 0, 0, 0, 1, 1, 0, 0, 0, 1, 0, 0, 1, 0, 1, 1], 
             'Time': ['00:28', '01:03', '01:51', '02:23', '03:23', '04:14', '05:08', '05:57', '06:33', '07:17', 
                      '08:15', '08:54', '09:48', '10:34', '11:28', '12:11', '12:46', '13:29', '14:14', '15:13', 
                      '16:06', '17:01', '17:42', '18:25', '19:22', '20:19', '20:51', '21:25', '22:07', '23:05']}, 
        '0003': 
            {
            'Stim': [0, 0, 1, 0, 0, 0, 0, 1, 0, 1, 0, 1, 0, 1, 1, 1, 0, 0, 1, 1, 0, 1, 1, 0, 1, 1, 0, 1, 1, 0], 
            'Time': ['00:29', '01:16', '01:46', '02:27', '03:10', '03:42', '04:18', '04:58', '05:31', '06:28', 
                     '07:27', '08:10', '08:41', '09:22', '10:17', '11:12', '12:03', '12:59', '13:35', '14:21', 
                     '15:12', '16:10', '16:51', '17:39', '18:13', '18:46', '19:35', '20:31', '21:21', '22:12']}, 

        '0004': 
            {'Stim': [1, 0, 0, 1, 1, 1, 1, 0, 1, 0, 1, 1, 1, 1, 1, 0, 0, 0, 0, 0, 1, 0, 0, 0, 0, 0, 1, 1, 1, 0], 
             'Time': ['00:30', '01:29', '02:07', '02:48', '03:39', '04:19', '04:59', '05:46', '06:20', '06:50', 
                      '07:49', '08:38', '09:37', '10:19', '11:16', '12:03', '12:42', '13:34', '14:08', '14:53', 
                      '15:25', '16:12', '16:56', '17:41', '18:11', '19:06', '19:43', '20:27', '21:02', '22:02']}, 
        '1193': 
            {'Stim': [1, 0, 1, 1, 0, 1, 0, 0, 1, 0, 1, 1, 0, 1, 1, 0, 0, 0, 1, 1, 0, 0, 0, 1, 0, 1, 0, 1, 1, 0] ,
            'Time': ['00:30', '01:25', '02:24', '03:12', '03:45', '04:17', '04:52', '05:38', '06:09', '06:44', 
                    '07:33', '08:17', '08:49', '09:42', '10:31', '11:30', '12:26', '13:23', '14:11', '15:08', 
                    '15:57', '16:30', '17:10', '18:06', '18:39', '19:33', '20:25', '21:21', '22:01', '22:34'] }, 
        '1194': 
            {'Stim': [0, 1, 0, 1, 0, 1, 0, 0, 1, 1, 1, 0, 0, 1, 0, 1, 0, 1, 0, 1, 1, 1, 0, 0, 0, 1, 0, 1, 0, 1], 
             'Time': ['00:30', '01:26', '02:22', '03:00', '03:46', '04:45', '05:44', '06:16', '06:47', '07:19', 
                    '08:07', '08:53', '09:39', '10:30', '11:21', '12:06', '12:38', '13:16', '14:08', '14:57', 
                    '15:48', '16:40', '17:26', '18:04', '18:56', '19:49', '20:29', '21:22', '22:22', '23:17']}, 
        '1195': 
            {'Stim': [0, 0, 1, 0, 1, 0, 0, 1, 1, 0, 0, 0, 1, 0, 1, 0, 1, 0, 1, 1, 0, 1, 0, 1, 0, 1, 1, 0, 1, 1],
             'Time': ['00:30', '01:09', '01:42', '02:21', '03:15', '04:06', '05:03', '05:47', '06:17', '07:13', 
                    '08:02', '08:32', '09:05', '09:44', '10:25', '11:03', '11:45', '12:54', '13:44',
                    '14:34', '15:05', '15:39', '16:14', '17:02', '17:37', '18:34', '19:14', '19:58', '20:52', '21:30']}, 
        '1196': {
            'Stim': [0, 1, 1, 1, 0, 1, 0, 1, 1, 0, 0, 1, 0, 1, 1, 0, 0, 0, 1, 1, 0, 1, 0, 1, 0, 0, 0, 1, 1, 0],
            'Time': ['00:33', '01:14', '01:55', '02:44', '03:23', '04:22', '05:10', '06:02', '06:55', '07:39', 
                    '08:34', '09:17', '10:08', '10:51', '11:45', '12:44', '13:28', '14:05', '15:00', '15:50', 
                    '16:35', '17:24', '18:09', '19:02', '19:39', '20:30', '21:06', '21:55', '22:31', '23:18']},

        '1198': 
            {
            'Stim': [1, 0, 0, 1, 0, 1, 1, 0, 0, 0, 1, 0, 0, 1, 1, 1, 0, 1, 0, 1, 0, 0, 0, 1, 1, 0, 1, 1, 1, 0],
            'Time': ['00:29', '01:23', '02:10', '03:00', '03:56', '04:34', '05:22', '06:06', '07:00', '07:41',
                     '08:20', '08:53', '09:44', '10:20', '11:20', '11:55', '12:41', '13:41', '14:19', '15:01',
                     '15:48', '16:26', '17:15', '17:48', '18:31', '19:27', '20:16', '20:50', '21:20', '22:13']}, 

        '1199': 
            {'Stim': [1, 1, 0, 0, 1, 1, 0, 1, 0, 1, 0, 0, 1, 0, 1, 0, 1, 1, 1, 0, 0, 0, 1, 0, 1, 1, 0, 0, 0, 1],
            'Time': ['00:29', '01:27', '02:07', '02:42', '03:27', '03:58', '04:32', '05:31', '06:10', '06:53',
                     '07:38', '08:24', '09:00', '09:50', '10:47', '11:42', '12:24', '13:08', '13:54', '14:53',
                     '15:35', '16:25', '17:13', '18:03', '18:34', '19:04', '19:51', '20:47', '21:18', '21:58']}, 

        '0011': 
            {'Stim': [1, 0, 1, 1, 0, 1, 0, 1, 1, 0, 0, 0, 1, 1, 1, 0, 0, 0, 1, 0, 0, 0, 1, 1, 1, 0, 1, 0, 0, 1], 
            'Time': ['00:29', '01:05', '01:42', '02:18', '03:12', '03:45', '04:45', '05:19', '06:01', '06:42', 
             '07:27', '08:27', '09:20', '10:12', '11:11', '12:00', '12:59', '13:33', '14:10', '14:47', 
             '15:25', '16:02', '16:50', '17:45', '18:37', '19:29', '20:26', '21:20', '22:08', '22:57']}, 

        '1200': 
            {'Stim': [0, 1, 0, 0, 1, 0, 0, 1, 1, 1, 0, 0, 1, 0, 1, 0, 0, 1, 1, 0, 0, 1, 0, 1, 1, 0, 1, 1, 0, 1], 
             'Time': ['00:30', '01:02', '01:47', '02:39', '03:22', '04:17', '05:04', '05:52', '06:39', '07:12', 
                    '07:42', '08:14', '09:14', '10:03', '10:39', '11:23', '12:15', '12:56', '13:52', '14:31', 
                    '15:07', '15:42', '16:37', '17:17', '18:15', '19:09', '20:08', '20:39', '21:19', '22:12']}, 

        '1201': 
            {'Stim': [1, 0, 1, 1, 0, 1, 0, 0, 1, 0, 1, 1, 0, 1, 1, 0, 0, 0, 1, 1, 0, 0, 1, 0, 1, 0, 1, 1, 0, 0,], 
             'Time': ['00:30', '01:09', '01:49', '02:29', '03:04', '04:01', '04:40', '05:25', '06:12', '06:48',
                     '07:34', '08:31', '09:10', '09:52', '10:43', '11:24', '12:00', '12:41', '13:13', '13:58', 
                     '14:53', '15:31', '17:06', '17:48', '18:23', '18:57', '19:56', '20:46', '21:18', '22:00']}}

trace_water_all = []
trace_tmt_all = []
trace_water_early_all = []
trace_tmt_early_all = []
trace_water_inter_all = []
trace_tmt_inter_all = []
trace_water_late_all = []
trace_tmt_late_all = []

    # Iterate over each PPD file in the directory
for ppd_file in ppd_files:
    ppd_file_path = os.path.join(ppd_files_dir, ppd_file)
    
    pre_start = 10
    #pre_start = 5
    post_start = 16
    
    # Extract the filename without the extension
    filename = os.path.splitext(os.path.basename(ppd_file_path))[0]
    
    mouse_number = filename[0:4]
    
    # Load the data from the CSV file
    data = import_ppd(ppd_file_path, low_pass=20, high_pass=0.001)
    
    # Convert sample index to time vector
    time = np.arange(len(data['analog_1'])) / 130
    
    # http://dx.doi.org/10.1016/j.cell.2015.07.014
    # dFF using 405 fit as baseline
    reg= np.polyfit(data['analog_2'], data['analog_1'], 1) # ch1 is 465nm, ch2 is 405nm 
    fit_405=reg[0]*data['analog_2']+reg[1]
    dFF=(data['analog_1']-fit_405)/fit_405 #this gives deltaF/F
    data['fit_405']=fit_405
    data['dFF']=dFF
    ch1=data['analog_1']
    ch2=data['analog_2']

    stim_data = timing_data[mouse_number]['Stim']
    time_stamps = timing_data[mouse_number]['Time']
    sampling_rate = 130
    
    def time_stamp_to_seconds(time_stamp):
        minutes, seconds = time_stamp.split(':')
        return int(minutes) * 60 + int(seconds)
    
    water_indexes = [i for i, stim in enumerate(stim_data) if stim == 0]
    tmt_indexes = [i for i, stim in enumerate(stim_data) if stim == 1]
    all_indexes = [i for i, stim in enumerate(stim_data)]
    
    water_time_stamps = [time_stamps[i] for i in water_indexes]
    tmt_time_stamps = [time_stamps[i] for i in tmt_indexes]
    all_time_stamps = [time_stamps[i] for i in all_indexes]
    
    water_data_indexes = [int(time_stamp_to_seconds(time_stamp) * sampling_rate) for time_stamp in water_time_stamps]
    tmt_data_indexes = [int(time_stamp_to_seconds(time_stamp) * sampling_rate) for time_stamp in tmt_time_stamps]
    all_data_indexes = [int(time_stamp_to_seconds(time_stamp) * sampling_rate) for time_stamp in all_time_stamps]
    
    # Create an empty matrix to store trace data
    trace_data_matrix = []
    
    
    for index in all_data_indexes:
        start = int(index - pre_start * sampling_rate)
        end = int(index + post_start * sampling_rate)
        trace_data = dFF[start:end]
        time = np.arange(start, end) / sampling_rate
    
        # Calculate the baseline value
        baseline_start = int(index - 5 * sampling_rate)
        baseline_end = int(index - 3 * sampling_rate)
        baseline_mean = np.mean(dFF[baseline_start:baseline_end])
        baseline_std = np.std(dFF[baseline_start:baseline_end])
    
        # Compute the relative trace data
        relative_trace_data = (trace_data - baseline_mean)# / baseline_std
        
        # Append relative_trace_data to the matrix
        trace_data_matrix.append(relative_trace_data)
    
    # Convert trace_data_matrix to a NumPy array
    trace_data_matrix = np.array(trace_data_matrix)
    trace_data_matrix_water = trace_data_matrix[water_indexes]
    trace_data_matrix_tmt = trace_data_matrix[tmt_indexes]
    
    # Calculate SEM for water and TMT
    sem_water = np.std(trace_data_matrix_water, axis=0) / np.sqrt(trace_data_matrix_water.shape[0])
    sem_tmt = np.std(trace_data_matrix_tmt, axis=0) / np.sqrt(trace_data_matrix_tmt.shape[0])
    
    # Calculate the average trace for water and TMT
    average_trace_water = np.mean(trace_data_matrix_water, axis=0)
    average_trace_tmt = np.mean(trace_data_matrix_tmt, axis=0)
    
    # Time array
    time_array = np.arange(-pre_start, post_start, 1/sampling_rate)
    
    plt.plot(time_array, average_trace_water * 100, label='Water', color='blue')
    plt.fill_between(time_array, (average_trace_water - sem_water) * 100, (average_trace_water + sem_water) * 100, color='blue', alpha=0.2)
    
    plt.plot(time_array, average_trace_tmt * 100, label='TMT', color='red')
    plt.fill_between(time_array, (average_trace_tmt - sem_tmt) * 100, (average_trace_tmt + sem_tmt) * 100, color='red', alpha=0.2)
    
    plt.axvline(x=0, color='k', linestyle='dashed')
    
    plt.xlabel('Time (seconds)')
    plt.ylabel('ΔF/F₀ (%)')
    plt.title('Mouse ' + mouse_number)
    plt.legend()
    # Save the figure as PNG with 300 dpi
    save_path = os.path.join(ppd_files_dir, mouse_number + 'avg.png')
    plt.savefig(save_path, dpi=300)
    plt.close()
    
    #Calculate the average trace for water and TMT
    average_trace_water_early = np.mean(trace_data_matrix_water[0:5,:], axis=0)
    average_trace_tmt_early = np.mean(trace_data_matrix_tmt[0:5,:], axis=0)
    average_trace_water_inter = np.mean(trace_data_matrix_water[5:10,:], axis=0)
    average_trace_tmt_inter = np.mean(trace_data_matrix_tmt[5:10,:], axis=0)
    average_trace_water_late = np.mean(trace_data_matrix_water[10:15,:], axis=0)
    average_trace_tmt_late = np.mean(trace_data_matrix_tmt[10:15,:], axis=0)
    
    trace_water_all.append(average_trace_water)
    trace_tmt_all.append(average_trace_tmt)
    trace_water_early_all.append(average_trace_water_early)
    trace_tmt_early_all.append(average_trace_tmt_early)
    trace_water_inter_all.append(average_trace_water_inter)
    trace_tmt_inter_all.append(average_trace_tmt_inter)
    trace_water_late_all.append(average_trace_water_late)
    trace_tmt_late_all.append(average_trace_tmt_late)
    
    plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), average_trace_water_early * 100, label='Water_early', color='blue')
    plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), average_trace_tmt_early * 100, label='TMT_early', color='red')
    plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), average_trace_water_inter * 100, label='Water_inter', color='cyan')
    plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), average_trace_tmt_inter * 100, label='TMT_inter', color='yellow')
    plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), average_trace_water_late * 100, label='Water_late', color='green')
    plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), average_trace_tmt_late * 100, label='TMT_late', color='magenta')
    plt.axvline(x=0, color='k', linestyle='dashed')
    plt.xlabel('Time (seconds)')
    plt.ylabel('ΔF/F₀ (%)')
    plt.title('Mouse ' + mouse_number)
    plt.legend()
    # Save the figure as PNG with 300 dpi
    save_path = os.path.join(ppd_files_dir, mouse_number + 'trials.png')
    plt.savefig(save_path, dpi=300)
    plt.close()
    
    # save mouse 521 tmt traces
    # if mouse_number == mouse_number: #str('521'):
    #     np.save('trace_tmt_' + mouse_number + '.npy', np.array(trace_data_matrix_tmt))
    #     np.save('trace_water_' + mouse_number + '.npy', np.array(trace_data_matrix_water))
    # Filter trace data for mouse 521
    tmt_trace_data = trace_data_matrix_tmt  # TMT trial traces
    water_trace_data = trace_data_matrix_water  # Water trial traces

    # Create the time vector for the first row
    pre_start = 10
    post_start = 16
    sampling_rate = 130  # Replace with your actual sampling rate
    time_vector = np.arange(-pre_start, post_start, 1/sampling_rate)

    # Convert TMT trace data to DataFrame and add time vector as the first row
    df_tmt = pd.DataFrame(tmt_trace_data)
    df_tmt.loc[-1] = time_vector  # Add time vector as the first row
    df_tmt.index = df_tmt.index + 1  # Shift index
    df_tmt.sort_index(inplace=True)  # Sort index

    # Convert Water trace data to DataFrame and add time vector as the first row
    df_water = pd.DataFrame(water_trace_data)
    df_water.loc[-1] = time_vector  # Add time vector as the first row
    df_water.index = df_water.index + 1  # Shift index
    df_water.sort_index(inplace=True)  # Sort index

    # Save to Excel
    tmt_excel_path = os.path.join(ppd_files_dir, f'Mouse_{mouse_number}_tmt_traces.xlsx')
    #df_tmt.to_excel(tmt_excel_path, index=False)

    water_excel_path = os.path.join(ppd_files_dir, f'Mouse_{mouse_number}_water_traces.xlsx')
    #df_water.to_excel(water_excel_path, index=False)

    #print(f'TMT traces saved to {tmt_excel_path}')
    #print(f'Water traces saved to {water_excel_path}')  

    smoothed_dFF = savgol_filter(dFF, window_length=51, polyorder=3)  

    # Create a dataframe to store the smoothed_dFF, time stamps, and stim data
    data_length = len(smoothed_dFF)
    stim_markers = np.full(data_length, -1)
    time = np.arange(0, len(smoothed_dFF)) / sampling_rate

    for i, index in enumerate(all_data_indexes):
        stim_markers[index] = stim_data[i]

    # Create a DataFrame
    df = pd.DataFrame({
        'Time (seconds)': time,
        'dFF': dFF,
        'smoothed_dFF': smoothed_dFF,
        'Stim': stim_markers
    })

    # Save the DataFrame to an Excel file
    water_excel_path = os.path.join(ppd_files_dir, f'Mouse_{mouse_number}_dFF_with_stim.xlsx')
    #df.to_excel(water_excel_path, index=False)

    # Convert timestamps to sample indices
    time_in_seconds = [time_stamp_to_seconds(ts) for ts in time_stamps]
    data_indexes = [int(ts * sampling_rate) for ts in time_in_seconds]

    # Set figure size to make the plot wider
    plt.figure(figsize=(20, 7))

    # Mark the timestamps in the trace array
    for i, stim in enumerate(stim_data):
        if stim == 0:
            plt.axvline(x=time_in_seconds[i], color='gray', linestyle='dashed', ymax=1)
        else:
            plt.axvline(x=time_in_seconds[i], color='gray', linestyle='solid', ymax=1)

    plt.plot(time, dFF, label='dFF')

    plt.xlabel('Time (seconds)')
    plt.ylabel('ΔF/F₀ (%)')
    plt.title('Mouse_' + mouse_number + ' Astrocyte activities during TMT and Water puff')
    #plt.savefig(os.path.join(ppd_files_dir, f'Mouse_{mouse_number}_Astrocyte_activities.png'), dpi=300)
    plt.close()

# Calculate the mean and SEM for each list
mean_trace_water = np.mean(trace_water_all, axis=0)
sem_trace_water = np.std(trace_water_all, axis=0) / np.sqrt(len(trace_water_all))

mean_trace_tmt = np.mean(trace_tmt_all, axis=0)
sem_trace_tmt = np.std(trace_tmt_all, axis=0) / np.sqrt(len(trace_tmt_all))

mean_trace_water_early = np.mean(trace_water_early_all, axis=0)
sem_trace_water_early = np.std(trace_water_early_all, axis=0) / np.sqrt(len(trace_water_early_all))

mean_trace_tmt_early = np.mean(trace_tmt_early_all, axis=0)
sem_trace_tmt_early = np.std(trace_tmt_early_all, axis=0) / np.sqrt(len(trace_tmt_early_all))

mean_trace_water_inter = np.mean(trace_water_inter_all, axis=0)
sem_trace_water_inter = np.std(trace_water_inter_all, axis=0) / np.sqrt(len(trace_water_inter_all))

mean_trace_tmt_inter = np.mean(trace_tmt_inter_all, axis=0)
sem_trace_tmt_inter = np.std(trace_tmt_inter_all, axis=0) / np.sqrt(len(trace_tmt_inter_all))

mean_trace_water_late = np.mean(trace_water_late_all, axis=0)
sem_trace_water_late = np.std(trace_water_late_all, axis=0) / np.sqrt(len(trace_water_late_all))

mean_trace_tmt_late = np.mean(trace_tmt_late_all, axis=0)
sem_trace_tmt_late = np.std(trace_tmt_late_all, axis=0) / np.sqrt(len(trace_tmt_late_all))


plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), np.mean(trace_tmt_all,axis=0) * 100, label='TMT', color='red')
plt.fill_between(np.arange(-pre_start, post_start, 1/sampling_rate), (mean_trace_tmt - sem_trace_tmt) * 100, (mean_trace_tmt + sem_trace_tmt) * 100, color='red', alpha=0.3)
plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), np.mean(trace_water_all,axis=0) * 100, label='Water', color='blue')
plt.fill_between(np.arange(-pre_start, post_start, 1/sampling_rate), (mean_trace_water - sem_trace_water) * 100, (mean_trace_water + sem_trace_water) * 100, color='blue', alpha=0.3)
#plt.ylim(-0.5, 3.5)
plt.axvline(x=0, color='k', linestyle='dashed')
plt.xlabel('Time from puffing (seconds)')
plt.ylabel('ΔF/F₀ (%)')
plt.title('Astrocyte Calcium traces for Water / TMT puffs')
plt.legend()
# Save the figure as PNG with 300 dpi
save_path = os.path.join(ppd_files_dir, 'tmt_water_overall.png')
plt.savefig(save_path, dpi=300)
plt.close() 

plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), np.mean(trace_tmt_early_all,axis=0) * 100, label='TMT', color='red')
plt.fill_between(np.arange(-pre_start, post_start, 1/sampling_rate), (mean_trace_tmt_early - sem_trace_tmt_early) * 100, (mean_trace_tmt_early + sem_trace_tmt_early) * 100, color='red', alpha=0.3)
plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), np.mean(trace_water_early_all,axis=0) * 100, label='Water', color='blue')
plt.fill_between(np.arange(-pre_start, post_start, 1/sampling_rate), (mean_trace_water_early - sem_trace_water_early) * 100, (mean_trace_water_early + sem_trace_water_early) * 100, color='blue', alpha=0.3)
#plt.ylim(-0.5, 3.5)
plt.axvline(x=0, color='k', linestyle='dashed')
plt.xlabel('Time from puffing (seconds)')
plt.ylabel('ΔF/F₀ (%)')
plt.title('Early Astrocyte Calcium traces for Water / TMT puffs')
plt.legend()
# Save the figure as PNG with 300 dpi
save_path = os.path.join(ppd_files_dir, 'tmt_water_early.png')
plt.savefig(save_path, dpi=300)
plt.close() 

plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), np.mean(trace_tmt_inter_all,axis=0) * 100, label='TMT', color='red')
plt.fill_between(np.arange(-pre_start, post_start, 1/sampling_rate), (mean_trace_tmt_inter - sem_trace_tmt_inter) * 100, (mean_trace_tmt_inter + sem_trace_tmt_inter) * 100, color='red', alpha=0.3)
plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), np.mean(trace_water_inter_all,axis=0) * 100, label='Water', color='blue')
plt.fill_between(np.arange(-pre_start, post_start, 1/sampling_rate), (mean_trace_water_inter - sem_trace_water_inter) * 100, (mean_trace_water_inter + sem_trace_water_inter) * 100, color='blue', alpha=0.3)
#plt.ylim(-0.5, 3.5)
plt.axvline(x=0, color='k', linestyle='dashed')
plt.xlabel('Time from puffing (seconds)')
plt.ylabel('ΔF/F₀ (%)')
plt.title('Intermediate GRABda traces for Water / TMT puffs')
plt.legend()
# Save the figure as PNG with 300 dpi
save_path = os.path.join(ppd_files_dir, 'tmt_water_inter.png')
plt.savefig(save_path, dpi=300)
plt.close() 

plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), np.mean(trace_tmt_late_all,axis=0) * 100, label='TMT', color='red')
plt.fill_between(np.arange(-pre_start, post_start, 1/sampling_rate), (mean_trace_tmt_late - sem_trace_tmt_late) * 100, (mean_trace_tmt_late + sem_trace_tmt_late) * 100, color='red', alpha=0.3)
plt.plot(np.arange(-pre_start, post_start, 1/sampling_rate), np.mean(trace_water_late_all,axis=0) * 100, label='Water', color='blue')
plt.fill_between(np.arange(-pre_start, post_start, 1/sampling_rate), (mean_trace_water_late - sem_trace_water_late) * 100, (mean_trace_water_late + sem_trace_water_late) * 100, color='blue', alpha=0.3)
#plt.ylim(-0.5, 3.5)
plt.axvline(x=0, color='k', linestyle='dashed')
plt.xlabel('Time from puffing (seconds)')
plt.ylabel('ΔF/F₀ (%)')
plt.title('Late GRABda traces for Water / TMT puffs')
plt.legend()
# Save the figure as PNG with 300 dpi
save_path = os.path.join(ppd_files_dir, 'tmt_water_late.png')
plt.savefig(save_path, dpi=300)
plt.close() 

# Define the new headers
headers = ['Mouse1193', 'Mouse1194', 'Mouse1195', 'Mouse1196', 'Mouse1198', 'Mouse1199', 'Mouse1200', 'Mouse1201']

# Save each array to an Excel file, transposed, and with new headers
pd.DataFrame(np.array(trace_water_all).T, columns=headers).to_excel('final_processed_trace_water_all.xlsx', index=False)
pd.DataFrame(np.array(trace_tmt_all).T, columns=headers).to_excel('final_processed_trace_tmt_all.xlsx', index=False)
pd.DataFrame(np.array(trace_water_early_all).T, columns=headers).to_excel('final_processed_trace_water_early_all.xlsx', index=False)
pd.DataFrame(np.array(trace_tmt_early_all).T, columns=headers).to_excel('final_processed_trace_tmt_early_all.xlsx', index=False)
pd.DataFrame(np.array(trace_water_inter_all).T, columns=headers).to_excel('final_processed_trace_water_inter_all.xlsx', index=False)
pd.DataFrame(np.array(trace_tmt_inter_all).T, columns=headers).to_excel('final_processed_trace_tmt_inter_all.xlsx', index=False)
pd.DataFrame(np.array(trace_water_late_all).T, columns=headers).to_excel('final_processed_trace_water_late_all.xlsx', index=False)
pd.DataFrame(np.array(trace_tmt_late_all).T, columns=headers).to_excel('final_processed_trace_tmt_late_all.xlsx', index=False)

# Create DataFrames
df_mean_trace_water = pd.DataFrame(np.array(mean_trace_water))
df_mean_trace_tmt = pd.DataFrame(np.array(mean_trace_tmt))
df_sem_trace_water = pd.DataFrame(np.array(sem_trace_water))
df_sem_trace_tmt = pd.DataFrame(np.array(sem_trace_tmt))

df_mean_trace_water_early = pd.DataFrame(np.array(mean_trace_water_early))
df_mean_trace_tmt_early = pd.DataFrame(np.array(mean_trace_tmt_early))
df_sem_trace_water_early = pd.DataFrame(np.array(sem_trace_water_early))
df_sem_trace_tmt_early = pd.DataFrame(np.array(sem_trace_tmt_early))

df_mean_trace_water_inter = pd.DataFrame(np.array(mean_trace_water_inter))
df_mean_trace_tmt_inter = pd.DataFrame(np.array(mean_trace_tmt_inter))
df_sem_trace_water_inter = pd.DataFrame(np.array(sem_trace_water_inter))
df_sem_trace_tmt_inter = pd.DataFrame(np.array(sem_trace_tmt_inter))

df_mean_trace_water_late = pd.DataFrame(np.array(mean_trace_water_late))
df_mean_trace_tmt_late = pd.DataFrame(np.array(mean_trace_tmt_late))
df_sem_trace_water_late = pd.DataFrame(np.array(sem_trace_water_late))
df_sem_trace_tmt_late = pd.DataFrame(np.array(sem_trace_tmt_late))

# Concatenate DataFrames side by side
df_combined = pd.concat([df_mean_trace_water, df_sem_trace_water, df_mean_trace_tmt, df_sem_trace_tmt, 
                         df_mean_trace_water_early, df_sem_trace_water_early, df_mean_trace_tmt_early, df_sem_trace_tmt_early,
                        df_mean_trace_water_inter, df_sem_trace_water_inter, df_mean_trace_tmt_inter, df_sem_trace_tmt_inter,
                        df_mean_trace_water_late, df_sem_trace_water_late, df_mean_trace_tmt_late, df_sem_trace_tmt_late
                         ], axis=1)

# Set column headings
df_combined.columns = ['mean_trace_water', 'sem_trace_water', 'mean_trace_tmt', 'sem_trace_tmt',
                       'mean_trace_water_early', 'sem_trace_water_early', 'mean_trace_tmt_early', 'sem_trace_tmt_early',
                       'mean_trace_water_inter', 'sem_trace_water_inter', 'mean_trace_tmt_inter', 'sem_trace_tmt_inter',
                       'mean_trace_water_late', 'sem_trace_water_late', 'mean_trace_tmt_late', 'sem_trace_tmt_late']

# Save to a single Excel file
df_combined.to_excel('combined_data.xlsx', index=False)

# Adjust column widths using openpyxl directly
wb = load_workbook('combined_data.xlsx')
ws = wb.active

# Define the column headers
headers = ['mean_trace_water', 'sem_trace_water', 'mean_trace_tmt', 'sem_trace_tmt',
            'mean_trace_water_early', 'sem_trace_water_early', 'mean_trace_tmt_early', 'sem_trace_tmt_early',
            'mean_trace_water_inter', 'sem_trace_water_inter', 'mean_trace_tmt_inter', 'sem_trace_tmt_inter',
            'mean_trace_water_late', 'sem_trace_water_late', 'mean_trace_tmt_late', 'sem_trace_tmt_late']

# Set the column width based on the length of the headers
for i, header in enumerate(headers, start=1):
    max_length = max(len(header), 10)  # Ensure a minimum width of 10 for better readability
    col_letter = ws.cell(row=1, column=i).column_letter
    ws.column_dimensions[col_letter].width = max_length

# Save the workbook
wb.save('grand_avg_traces.xlsx')


np.save('trace_water_all.npy', np.array(trace_water_all))
np.save('trace_tmt_all.npy', np.array(trace_tmt_all))
np.save('trace_water_early_all.npy', np.array(trace_water_early_all))
np.save('trace_tmt_early_all.npy', np.array(trace_tmt_early_all))
np.save('trace_water_inter_all.npy', np.array(trace_water_inter_all))
np.save('trace_tmt_inter_all.npy', np.array(trace_tmt_inter_all))
np.save('trace_water_late_all.npy', np.array(trace_water_late_all))
np.save('trace_tmt_late_all.npy', np.array(trace_tmt_late_all))



import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split, cross_val_score
from sklearn.svm import SVC
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import classification_report, accuracy_score

# Prepare the data
# Assuming trace_tmt_all and trace_water_all are the dFF signals for TMT and Water
# Combine and label the data
X = np.concatenate((trace_tmt_all, trace_water_all), axis=0)
y = np.array([1] * len(trace_tmt_all) + [0] * len(trace_water_all))

# Flatten the data for SVM input
X = X.reshape(X.shape[0], -1)

# Standardize the features
scaler = StandardScaler()
X = scaler.fit_transform(X)

# Split the data into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)

# Train the SVM model
svm_model = SVC(kernel='linear', C=1)
svm_model.fit(X_train, y_train)

# Evaluate the model
y_pred = svm_model.predict(X_test)
print("SVM Classification Report:")
print(classification_report(y_test, y_pred))
print("Accuracy:", accuracy_score(y_test, y_pred))

# Cross-validation
cv_scores = cross_val_score(svm_model, X, y, cv=5)
print("Cross-validation Accuracy Scores:", cv_scores)
print("Mean Cross-validation Accuracy:", cv_scores.mean())
